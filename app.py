import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import json

<<<<<<< HEAD
class ExcelViewerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Viewer")
        self.excel_files = []
        self.selected_sheets = []
        self.selected_sheet = tk.StringVar()
        self.selected_lab = tk.StringVar()
        self.selected_categoria = tk.StringVar()
        

        # Laboratorio, hoja y fechas de filtro
        self.label_laboratorio = tk.Label(self.root, text="Laboratorio/Droguería:")
        self.label_laboratorio.grid(row=0, column=0)
        self.entry_laboratorio = AutocompleteCombobox(self.root, textvariable=self.selected_lab, width=27)
        self.entry_laboratorio.grid(row=0, column=1)

        # Botón para seleccionar categoría
        self.label_categoria = tk.Label(self.root, text="Categoría:")
        self.label_categoria.grid(row=0, column=2)
        self.entry_categoria = AutocompleteCombobox(self.root, textvariable=self.selected_categoria, width=27)
        self.entry_categoria.grid(row=0, column=3)
        self.entry_categoria.set_completion_list(['No medicinal', 'Medicinal'])

        self.label_sheet = tk.Label(self.root, text="Hoja:")
        self.label_sheet.grid(row=1, column=0)
        self.sheet_dropdown = ttk.Combobox(self.root, textvariable=self.selected_sheet, state="readonly", width=30)
        self.sheet_dropdown.grid(row=1, column=1)
=======
archivos_cargados = []

def buscar_archivos():
    directorio = filedialog.askdirectory()
    archivos = [os.path.join(directorio, archivo) for archivo in os.listdir(directorio) if archivo.endswith('.xlsx') and not archivo.startswith('~$')]
    for archivo in archivos:
        print(archivo)
    return archivos

def cargar_datos():
    global archivos_cargados
    archivos_cargados = buscar_archivos()
    for archivo in archivos_cargados:
        df = pd.read_excel(archivo)

def filtrar_datos(archivos_cargados, laboratorio, fecha_inicio, fecha_fin):
    datos_totales = []
    # Iterar sobre cada archivo cargado
    for archivo in archivos_cargados:
        # Leer el archivo
        df = pd.read_excel(archivo)
        # Convertir la columna 'Laboratorio' a cadena (str) si no lo es
        df['Laboratorio'] = df['Laboratorio'].astype(str)
        # Convertir la columna 'Fecha' a formato de fecha y hora (datetime) si no lo es
        df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce')
        # Filtrar por laboratorio y rango de fechas
        filtro = (df['Laboratorio'] == laboratorio) & (df['Fecha'].between(fecha_inicio, fecha_fin))

        datos_filtrados = df[filtro]
        datos_totales.append(datos_filtrados)
    # Concatenar los datos filtrados en un único DataFrame
    return pd.concat(datos_totales)
>>>>>>> 0bbd141a1a6aa97d44678207b1b61cfb4798c391

        self.label_fecha_inicio = tk.Label(self.root, text="Fecha inicio (YYYY-MM-DD):")
        self.label_fecha_inicio.grid(row=2, column=0)
        self.entry_fecha_inicio = tk.Entry(self.root)
        self.entry_fecha_inicio.grid(row=2, column=1)

        self.label_fecha_fin = tk.Label(self.root, text="Fecha fin (YYYY-MM-DD):")
        self.label_fecha_fin.grid(row=3, column=0)
        self.entry_fecha_fin = tk.Entry(self.root)
        self.entry_fecha_fin.grid(row=3, column=1)

        # Botón para cargar archivos
        self.btn_cargar = tk.Button(self.root, text="Cargar archivos", command=self.load_files)
        self.btn_cargar.grid(row=4, column=0, columnspan=2, pady=10)

        # Botón para previsualizar los datos
        self.btn_previsualizar = tk.Button(self.root, text="Previsualizar datos", command=self.preview_sheet)
        self.btn_previsualizar.grid(row=5, column=0, columnspan=2, pady=10)

        # Texto para mostrar previsualización
        self.previsualizacion_texto = tk.Text(self.root)
        self.previsualizacion_texto.grid(row=6, column=0, columnspan=2, padx=10, pady=10)

        # Cargar los laboratorios desde el archivo Excel
        self.load_laboratorios()

    def load_laboratorios(self):
        # Ruta al archivo JSON
        json_file_path = os.path.join(os.getcwd(), "Convertir.json")
        try:
            with open(json_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                print(data)  # Imprimir el contenido del JSON para depurar
                laboratorios = [item["Laboratorio"] for item in data]
                self.entry_laboratorio.set_completion_list(laboratorios)
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar los laboratorios desde el archivo JSON: {str(e)}")
            print(e)
            
    def load_files(self):
        file_paths = filedialog.askopenfilenames(filetypes=[("Archivos de Excel", "*.xlsx")])
        if file_paths:
            self.excel_files.extend(file_paths)
            for file_path in file_paths:
                self.load_sheets(file_path)


    def load_sheets(self, file_path):
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            self.available_sheets = wb.sheetnames
            self.sheet_dropdown["values"] = self.available_sheets
        except Exception as e:
            messagebox.showwarning("Advertencia", f"Error al abrir el archivo {file_path}: {e}")


    def preview_sheet(self):
        laboratorio = self.selected_lab.get()
        sheet_name = self.selected_sheet.get()
        fecha_inicio = self.entry_fecha_inicio.get()
        fecha_fin = self.entry_fecha_fin.get()
        categoria = self.selected_categoria.get()
        if not laboratorio or not sheet_name or not fecha_inicio or not fecha_fin or not categoria:
            messagebox.showerror("Error", "Por favor ingresa el laboratorio, la hoja, las fechas y la categoría.")
            return

        datos_filtrados = self.filtrar_datos(laboratorio, sheet_name, fecha_inicio, fecha_fin, categoria)

        if datos_filtrados.empty:
            messagebox.showinfo("Información", "No hay datos disponibles para el filtro especificado.")
        else:
            self.previsualizacion_texto.delete(1.0, tk.END)
            self.previsualizacion_texto.insert(tk.END, datos_filtrados.to_string(index=False))

    def filtrar_datos(self, laboratorio, sheet_name, fecha_inicio, fecha_fin, categoria):
        datos_totales = []
        for file_path in self.excel_files:
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
            except Exception as e:
                messagebox.showwarning("Advertencia", f"Error al abrir el archivo {file_path}: {e}")
                continue
            
            # Verificar si la hoja seleccionada por el usuario está disponible en el archivo Excel
            if sheet_name not in wb.sheetnames:
                messagebox.showwarning("Advertencia", f"La hoja '{sheet_name}' no está disponible en el archivo {file_path}.")
                continue

            # Buscar la celda que contiene el texto "Laboratorio/Droguería" en la hoja seleccionada
            sheet = wb[sheet_name]
            lab_cell = None
            for row in sheet.iter_rows():
                for cell in row:
                    if "Laboratorio/Droguería" in cell.value.upper():
                        lab_cell = str(cell)
                        break
                if lab_cell:
                    break
            
            if not lab_cell:
                messagebox.showwarning("Advertencia", f"No se encontró la celda con 'Laboratorio/Droguería' en la hoja '{sheet_name}' del archivo {file_path}.")
                continue

            # Obtener el nombre del laboratorio en la siguiente fila y en la misma columna
            lab_row = lab_cell.row + 1
            lab_col = lab_cell.column
            lab_name = sheet.cell(row=lab_row, column=lab_col).value

            # Convertir el nombre del laboratorio y el valor proporcionado por el usuario a mayúsculas para comparar
            lab_name_upper = lab_name.upper()
            laboratorio_upper = laboratorio.upper()

            # Comparar los nombres de laboratorio
            if lab_name_upper != laboratorio_upper:
                messagebox.showwarning("Advertencia", f"El laboratorio '{laboratorio}' no coincide con el encontrado en la hoja '{sheet_name}' del archivo {file_path}.")
                continue

            # Filtrar los datos según los criterios en la hoja seleccionada
            datos_filtrados = []
            for row in sheet.iter_rows(min_row=lab_row + 1):
                values = [cell.value for cell in row]
                if len(values) < 4:  # Ignorar las filas con menos de 4 columnas (no tienen suficientes datos)
                    continue
                fecha_pedido = values[2]
                if fecha_pedido and fecha_inicio <= fecha_pedido <= fecha_fin:
                    if categoria == "Medicinal" and values[3] == "Medicinal":
                        datos_filtrados.append(values)
                    elif categoria == "No Medicinal" and values[3] == "No Medicinal":
                        datos_filtrados.append(values)
            if datos_filtrados:
                datos_totales.append(datos_filtrados)

        if datos_totales:
            # Convertir los datos a un DataFrame de pandas
            flat_data = [item for sublist in datos_totales for item in sublist]
            df = pd.DataFrame(flat_data)
            df.columns = [cell.value for cell in next(sheet.iter_rows())]  # Asignar nombres de columnas desde la primera fila
            return df
        else:
            return pd.DataFrame()



class AutocompleteCombobox(ttk.Combobox):
    def set_completion_list(self, completion_list):
        self._completion_list = sorted(completion_list, key=str.lower)  # Case-insensitive sorting
        self._hits = []
        self._hit_index = 0
        self.position = 0
        self.bind('<KeyRelease>', self.handle_keyrelease)
        self['values'] = self._completion_list

    def autocomplete(self, delta=0):
        if delta:
            self.delete(self.position, tk.END)
        else:
            self.position = len(self.get())
        _hits = []
        for item in self._completion_list:
            if item.lower().startswith(self.get().lower()):
                _hits.append(item)
        self._hits = _hits
        if _hits:
            self._hit_index = 0
            self.delete(0, tk.END)
            self.insert(0, _hits[0])
            self.select_range(self.position, tk.END)

    def handle_keyrelease(self, event):
        if event.keysym == "BackSpace":
            self.delete(self.index(tk.INSERT), tk.END)
            self.position = self.index(tk.END)
        if event.keysym == "Left":
            if self.position < self.index(tk.END):
                self.delete(self.position, tk.END)
            else:
                self.position = self.position - 1
                self.delete(self.position, tk.END)
        if event.keysym == "Right":
            self.position = self.position + 1
        if len(event.keysym) == 1:
            self.autocomplete()
        if event.keysym == "Down":
            self.autocomplete(1)
        if event.keysym == "Up":
            self.autocomplete(-1)

def main():
    root = tk.Tk()
    app = ExcelViewerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
