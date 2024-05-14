import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pandastable import Table
import openpyxl
from tkcalendar import DateEntry
import json
from ttkwidgets.autocomplete import AutocompleteCombobox
from datetime import datetime
import locale
from tqdm import tqdm
from tkinter.ttk import Progressbar

class SumarizadorNotasPedido:
    def __init__(self, root):
        self.root = root
        self.root.title("Sumarizador de Notas de Pedido")
        self.root.geometry("600x300")
        self.root.iconbitmap("images/descarga.ico")

        self.progress_label = None
        

        style = ttk.Style()
        style.theme_use('clam')  

        # Personalizar el estilo de la ventana y los widgets
        style.configure('.', background='#E0FFFF')  # Establecer un color de fondo celeste agua pastel
        style.configure('TLabel', font=('Arial', 12))  # Personalizar la fuente de las etiquetas
        style.configure('TButton', font=('Arial', 12), foreground='black', background='#87CEEB')
        

        self.df = None
        self.df_sumarizado = None
        
        # Cargar laboratorios desde archivo JSON
        with open("laboratorios.json", encoding='utf-8') as f:
            data = json.load(f)
            self.laboratorios = sorted(set(d["Laboratorio"] for d in data))  # Ordenar alfabéticamente
            
        with open("laboratoriosNuevo", 'w', encoding='utf-8') as f:
            json.dump(self.laboratorios, f, ensure_ascii=False, indent=4)

        
        #Interfaz grafica
        self.btn_cargar = tk.Button(self.root, text="Cargar archivos", command=self.cargar_archivos, borderwidth=2, relief = "raised")
        self.btn_cargar.pack(pady=10)

        

        self.lbl_fecha_inicio = tk.Label(self.root, text="Seleccione fecha de inicio:")
        self.lbl_fecha_inicio.pack()
        self.cal_fecha_inicio = DateEntry(self.root, width=12, background='darkblue', foreground='white', borderwidth=2, locale='es_ES', relief = "raised")
        self.cal_fecha_inicio.pack()

        self.lbl_fecha_fin = tk.Label(self.root, text="Seleccione fecha de fin:")
        self.lbl_fecha_fin.pack()
        self.cal_fecha_fin = DateEntry(self.root, width=12, background='darkblue', foreground='white', borderwidth=2, locale='es_ES', relief = "raised")
        self.cal_fecha_fin.pack()

        
        self.lbl_laboratorio = tk.Label(self.root, text="Seleccione el laboratorio:")
        self.lbl_laboratorio.pack()
        self.selected_laboratorio = tk.StringVar()
        self.entry_laboratorio = AutocompleteCombobox(self.root, textvariable=self.selected_laboratorio, completevalues=self.laboratorios)
        self.entry_laboratorio.pack()

        
        self.frame_botones = tk.Frame(self.root)
        self.frame_botones.pack(pady=10)
        
        self.btn_sumarizar = tk.Button(self.frame_botones, text="Sumarizar", command=self.sumarizar, borderwidth=2, relief = "raised")
        self.btn_sumarizar.pack(side=tk.LEFT, padx=5)
        
        self.btn_descargar = tk.Button(self.frame_botones, text="Descargar Resultados", command=self.descargar_resultados, borderwidth=2, relief = "raised")
        self.btn_descargar.pack(side=tk.LEFT, padx=5)

        
        self.lbl_resultado = tk.Label(self.root, text="")
        self.lbl_resultado.pack(pady=10)

    def cargar_archivos(self):
        files_selected = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
        if files_selected:
            try:
                dfs = []  # Array para almacenar los DataFrames cargados de los archivos

                # Configurar la barra de progreso
                progress_bar = ttk.Progressbar(self.root, orient="horizontal", length=300, mode="determinate")
                progress_bar.pack()

                num_files = len(files_selected)
                progress_bar['maximum'] = num_files

                # Configurar etiqueta para mostrar el porcentaje de progreso
                self.progress_label = tk.Label(self.root, text="")
                self.progress_label.pack()

                for idx, file_path in enumerate(files_selected, start=1):
                    print("Procesando archivo:", file_path)  # Depuración

                    try:
                        # Leer todas las hojas del archivo Excel
                        xl = pd.ExcelFile(file_path)
                        for sheet_name in xl.sheet_names:
                            print("Procesando hoja:", sheet_name)  # Depuración
                            try:
                                df = pd.read_excel(xl, sheet_name=sheet_name, skiprows=9)

                                # ... Código de procesamiento adicional ...

                                if "Can" not in df.columns or "Imp. Total" not in df.columns:
                                    print(f"No se encontraron las columnas necesarias en la hoja '{sheet_name}' del archivo '{file_path}'. Saltando esta hoja...")  # Mensaje de depuración
                                    continue

                                # Reemplaza los valores vacíos en la columna 'Codebar' con una cadena vacía
                                df["Codebar"] = df["Codebar"].fillna('').astype(str)

                                wb = openpyxl.load_workbook(file_path)
                                ws = wb[sheet_name]
                                fecha_pedido = ws['I2'].value 

                                df["Fecha del pedido"] = pd.to_datetime(fecha_pedido, errors='coerce')

                                # Asignación de celda fija para dataFrame

                                laboratorio = ws['E2'].value
                                drogueria = ws['E4'].value if ws['E4'].value else "Comprador no asignado"
                                comprador = ws['C6'].value if ws['C6'].value else "Comprador no asignado"

                                # Asignar Dataframe
                                df["Laboratorio"] = laboratorio
                                df["Droguería por donde Llega"] = drogueria
                                df["Comprador"] = comprador

                                dfs.append(df)
                            except Exception as e:
                                print(f"Error al procesar la hoja '{sheet_name}' del archivo '{file_path}': {e}")  # Depuración
                    except Exception as e:
                        print(f"Error al procesar el archivo {file_path}: {e}")  # Depuración

                    # Actualizar la barra de progreso
                    progress_bar['value'] = idx
                    self.progress_label.config(text=f"{round(idx/num_files*100)}% completado")  # Actualizar etiqueta de progreso
                    progress_bar.update_idletasks()

                if not dfs:
                    messagebox.showerror("Error", "No se encontraron archivos Excel seleccionados.")
                    return

                self.df = pd.concat(dfs)

                messagebox.showinfo("Éxito", "Archivos cargados correctamente.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudieron cargar los archivos: {e}")

            # Ocultar la barra de progreso y la etiqueta después de completar la carga
            progress_bar.pack_forget()
            self.progress_label.pack_forget()

    def sumarizar(self):
        if self.df is None:
            messagebox.showerror("Error", "Primero carga un archivo.")
            return

        try:
            self.df["Can"] = pd.to_numeric(self.df["Can"], errors="coerce")
            self.df["Imp. Total"] = pd.to_numeric(self.df["Imp. Total"], errors="coerce")

            fecha_inicio = self.cal_fecha_inicio.get_date()  
            fecha_fin = self.cal_fecha_fin.get_date()  
            laboratorio = self.selected_laboratorio.get()

            fecha_inicio = datetime.combine(fecha_inicio, datetime.min.time())
            fecha_fin = datetime.combine(fecha_fin, datetime.max.time())

            df_filtrado = self.df[
                (self.df["Fecha del pedido"] >= fecha_inicio) &
                (self.df["Fecha del pedido"] <= fecha_fin) &
                (self.df["Laboratorio"].str.lower().str.contains(laboratorio.lower()))
            ]

            df_filtrado = df_filtrado.dropna(subset=["Imp. Total"])
            df_filtrado = df_filtrado[df_filtrado["Imp. Total"] != 0]

            if df_filtrado.empty:
                messagebox.showerror("Error", "No se encontraron datos para el rango de fechas y laboratorio ingresados.")
                return

            resumen = df_filtrado.groupby(["Codebar", "Producto", "Droguería por donde Llega", "Comprador"]).agg(
                {'Can': 'sum', 'Imp. Total': 'sum'})

            resumen = resumen[resumen["Imp. Total"] != 0]

            total_cantidades = resumen["Can"].sum()
            total_importes = resumen["Imp. Total"].sum()

            totales = pd.DataFrame({"Codebar": ["TOTAL"], "Producto": [None], "Droguería por donde Llega": [None],
                            "Comprador": [None], "Can": [total_cantidades], "Imp. Total": [total_importes]})

            resumen_con_total = pd.concat([resumen.reset_index(), totales], ignore_index=True)

            resumen_con_total['Codebar'] = resumen_con_total['Codebar'].astype(str).str.replace('.0', '')

            self.df_sumarizado = resumen_con_total

            self.mostrar_resultado(resumen_con_total)

            messagebox.showinfo("Éxito", "Resultados sumarizados correctamente.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo sumarizar los datos: {e}")

    def mostrar_resultado(self, df):
        
        self.top = tk.Toplevel(self.root)
        self.top.title("Resultados Sumarizados")

        
        pt = Table(self.top, dataframe=df)
        pt.show()

    def descargar_resultados(self):
        if self.df_sumarizado is None:
            messagebox.showerror("Error", "No hay datos sumarizados para descargar.")
            return

        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
        if filename:
            try:
                with pd.ExcelWriter(filename) as writer:
                    self.df_sumarizado.to_excel(writer, index=False, sheet_name="Datos Sumarizados")
                    workbook = writer.book
                    worksheet = writer.sheets["Datos Sumarizados"]
                    header_format = workbook.add_format({'bold': True, 'align': 'center'})
                    worksheet.write(0, 0, "Codebar", header_format)
                    worksheet.write(0, 1, "Producto", header_format)
                    worksheet.write(0, 2, "Droguería por donde Llega", header_format)
                    worksheet.write(0, 3, "Comprador", header_format)
                    worksheet.write(0, 4, "Cantidad", header_format)
                    worksheet.write(0, 5, "Imp. Total", header_format)
                    worksheet.set_column('A:G', 20)
                    messagebox.showinfo("Éxito", "Resultados sumarizados descargados correctamente.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo descargar los resultados sumarizados: {e}")


locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')


root = tk.Tk()
app = SumarizadorNotasPedido(root)
root.mainloop()